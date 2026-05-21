from __future__ import annotations

import io
from pathlib import PurePosixPath

import openpyxl
import pytest

from docupipe.models import Bundle, FileItem
from docupipe.steps.excel_structured import ExcelStructuredStep


def _make_xlsx_bytes(sheets: dict[str, list[list]]) -> bytes:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for name, rows in sheets.items():
        ws = wb.create_sheet(name)
        for row in rows:
            ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _make_xlsx_bundle(sheets: dict[str, list[list]], **extra_ctx) -> Bundle:
    return Bundle(
        files=[FileItem(name="test.xlsx", content=_make_xlsx_bytes(sheets), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", role="main")],
        context={"extension": "xlsx", "id": "doc1", "title": "test.xlsx", "path": "test.xlsx", **extra_ctx},
    )


class TestExcelStructuredPassthrough:
    def test_non_xlsx_extension_returns_unchanged(self):
        bundle = Bundle(
            files=[FileItem(name="doc.pdf", content=b"pdf-data", content_type="application/pdf", role="main")],
            context={"extension": "pdf"},
        )
        step = ExcelStructuredStep()
        result = step.process(bundle)
        assert result is bundle
        assert result.main.content == b"pdf-data"

    def test_no_extension_returns_unchanged(self):
        bundle = Bundle(
            files=[FileItem(name="doc", content=b"data", content_type="", role="main")],
            context={},
        )
        step = ExcelStructuredStep()
        result = step.process(bundle)
        assert result is bundle

    def test_no_main_file_returns_unchanged(self):
        bundle = Bundle(files=[], context={"extension": "xlsx"})
        step = ExcelStructuredStep()
        result = step.process(bundle)
        assert result is bundle


class TestExcelStructuredSingleSheet:
    def test_single_sheet_produces_one_main_file(self):
        bundle = _make_xlsx_bundle({"员工表": [["姓名", "部门"], ["张三", "销售部"]]})
        step = ExcelStructuredStep()
        result = step.process(bundle)

        assert len(result.files) == 1
        main = result.main
        assert main is not None
        assert main.role == "main"
        assert main.content_type == "text/markdown"
        assert main.name == "test_员工表.md"
        assert "张三" in main.content
        assert "销售部" in main.content
        assert "姓名" in main.content
        assert "部门" in main.content

    def test_extension_updated_to_md(self):
        bundle = _make_xlsx_bundle({"Sheet1": [["A", "B"], ["1", "2"]]})
        step = ExcelStructuredStep()
        result = step.process(bundle)
        assert result.context["extension"] == "md"

    def test_markdown_table_format(self):
        bundle = _make_xlsx_bundle({"S": [["姓名", "年龄"], ["张三", "30"]]})
        step = ExcelStructuredStep()
        result = step.process(bundle)
        lines = result.main.content.strip().split("\n")
        assert lines[0] == "## S"
        assert lines[1] == ""
        assert lines[2] == "| 姓名 | 年龄 |"
        assert "|---" in lines[3]
        assert lines[4] == "| 张三 | 30 |"


class TestExcelStructuredMultipleSheets:
    def test_two_sheets_produces_main_and_attachment(self):
        bundle = _make_xlsx_bundle({
            "Sheet1": [["A"], ["1"]],
            "Sheet2": [["B"], ["2"]],
        })
        step = ExcelStructuredStep()
        result = step.process(bundle)

        assert len(result.files) == 2
        assert result.files[0].role == "main"
        assert result.files[1].role == "attachment"
        assert "Sheet1" in result.files[0].name
        assert "Sheet2" in result.files[1].name

    def test_three_sheets_produces_one_main_two_attachments(self):
        bundle = _make_xlsx_bundle({
            "A": [["x"], ["1"]],
            "B": [["y"], ["2"]],
            "C": [["z"], ["3"]],
        })
        step = ExcelStructuredStep()
        result = step.process(bundle)

        assert len(result.files) == 3
        roles = [f.role for f in result.files]
        assert roles == ["main", "attachment", "attachment"]


class TestExcelStructuredMergedCells:
    def test_merged_cells_filled(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Merged"
        ws["A1"] = "合并值"
        ws.merge_cells("A1:B1")
        ws["A2"] = "其他"
        ws["B2"] = "数据"

        buf = io.BytesIO()
        wb.save(buf)
        bundle = Bundle(
            files=[FileItem(name="merged.xlsx", content=buf.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", role="main")],
            context={"extension": "xlsx", "id": "d1", "title": "merged.xlsx", "path": "merged.xlsx"},
        )

        step = ExcelStructuredStep(fill_merged=True)
        result = step.process(bundle)
        content = result.main.content
        # 合并单元格填充后，A1 和 B1 都应该有值
        assert "合并值" in content

    def test_merged_cells_not_filled_when_disabled(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Merged"
        ws["A1"] = "合并值"
        ws.merge_cells("A1:B1")
        ws["A2"] = "其他"
        ws["B2"] = "数据"

        buf = io.BytesIO()
        wb.save(buf)
        bundle = Bundle(
            files=[FileItem(name="merged.xlsx", content=buf.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", role="main")],
            context={"extension": "xlsx", "id": "d1", "title": "merged.xlsx", "path": "merged.xlsx"},
        )

        step = ExcelStructuredStep(fill_merged=False)
        result = step.process(bundle)
        # fill_merged=False 仍然处理文件，只是不填充合并单元格
        assert result.main is not None
        assert "合并值" in result.main.content
        assert "其他" in result.main.content


class TestExcelStructuredHiddenRows:
    def test_hidden_rows_skipped(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Hidden"
        ws.append(["姓名", "部门"])
        ws.append(["张三", "销售部"])
        ws.append(["李四", "技术部"])
        ws.row_dimensions[3].hidden = True
        ws.append(["王五", "市场部"])

        buf = io.BytesIO()
        wb.save(buf)
        bundle = Bundle(
            files=[FileItem(name="hidden.xlsx", content=buf.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", role="main")],
            context={"extension": "xlsx"},
        )

        step = ExcelStructuredStep(skip_hidden=True)
        result = step.process(bundle)
        content = result.main.content
        assert "张三" in content
        assert "王五" in content
        assert "李四" not in content


class TestExcelStructuredHiddenCols:
    def test_hidden_columns_skipped(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "HiddenCol"
        ws.append(["姓名", "秘密", "部门"])
        ws.column_dimensions["B"].hidden = True
        ws.append(["张三", "机密", "销售部"])

        buf = io.BytesIO()
        wb.save(buf)
        bundle = Bundle(
            files=[FileItem(name="hidden_col.xlsx", content=buf.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", role="main")],
            context={"extension": "xlsx"},
        )

        step = ExcelStructuredStep(skip_hidden=True)
        result = step.process(bundle)
        content = result.main.content
        assert "姓名" in content
        assert "部门" in content
        assert "秘密" not in content
        assert "机密" not in content


class TestExcelStructuredEmptyRowsCols:
    def test_empty_rows_skipped(self):
        bundle = _make_xlsx_bundle({"S": [["A"], ["1"], [], ["2"]]})
        step = ExcelStructuredStep(skip_empty=True)
        result = step.process(bundle)
        content = result.main.content
        assert "1" in content
        assert "2" in content
        data_lines = [l for l in content.split("\n") if l.strip().startswith("|") and not l.startswith("|-")]
        assert len(data_lines) == 3  # header + 2 data rows

    def test_empty_rows_kept_when_disabled(self):
        bundle = _make_xlsx_bundle({"S": [["A"], ["1"], [], ["2"]]})
        step = ExcelStructuredStep(skip_empty=False)
        result = step.process(bundle)
        content = result.main.content
        data_lines = [l for l in content.split("\n") if l.strip().startswith("|") and not l.startswith("|-")]
        assert len(data_lines) >= 3


class TestExcelStructuredIntegration:
    def test_full_pipeline_excel_to_hindsight_items(self):
        """模拟完整 pipeline: excel_structured → HindsightDestination._build_retain_item"""
        from docupipe.config import resolve_context_vars
        from docupipe.destinations.hindsight import HindsightDestination

        bundle = _make_xlsx_bundle(
            {"员工": [["姓名", "部门"], ["张三", "销售部"]], "项目": [["名称", "状态"], ["Alpha", "进行中"]]},
            _source="local",
            hash="abc123",
        )

        step = ExcelStructuredStep()
        result = step.process(bundle)

        assert len(result.files) == 2
        assert result.files[0].role == "main"
        assert result.files[1].role == "attachment"

        dest = HindsightDestination(
            bank_id="test",
            api_url="http://localhost",
            api_key="k",
            document_id_template="${context._source}:${context.id}:${context._sheet_name}",
        )
        dest._process_roles = ["main", "attachment"]

        for file_item in result.files:
            sheet_name = PurePosixPath(file_item.name).stem
            # 模拟 pipeline 层的 resolve_context_vars + update_config
            resolved_config = resolve_context_vars(
                {"document_id_template": dest._document_id_template},
                {**result.context, "_sheet_name": sheet_name},
            )
            dest.update_config(resolved_config)
            item = dest._build_retain_item(result, file_item=file_item, sheet_name=sheet_name)
            assert item["content"]
            assert "local:doc1:" in item["document_id"]
