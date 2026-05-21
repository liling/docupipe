from __future__ import annotations

import io
import logging
from pathlib import PurePosixPath

import openpyxl
from openpyxl.utils import column_index_from_string

from docupipe.models import Bundle, FileItem
from docupipe.steps import register_step
from docupipe.steps.base import Step

logger = logging.getLogger(__name__)


@register_step("excel_structured")
class ExcelStructuredStep(Step):
    _config_keys = {"fill_merged", "skip_hidden", "skip_empty"}

    def __init__(
        self,
        fill_merged: bool = True,
        skip_hidden: bool = True,
        skip_empty: bool = True,
        **kwargs,
    ):
        self._fill_merged = fill_merged
        self._skip_hidden = skip_hidden
        self._skip_empty = skip_empty

    def process(self, bundle: Bundle) -> Bundle:
        ext = bundle.context.get("extension", "")
        if ext != "xlsx":
            return bundle

        main = bundle.main
        if not main or not isinstance(main.content, bytes):
            return bundle

        wb = openpyxl.load_workbook(io.BytesIO(main.content), read_only=False, data_only=True)

        if self._fill_merged:
            self._fill_merged_cells(wb)

        stem = PurePosixPath(main.name).stem
        new_files: list[FileItem] = []

        for ws in wb.worksheets:
            md = self._sheet_to_markdown(ws)
            if not md.strip():
                continue

            role = "main" if not new_files else "attachment"
            sheet_name = ws.title
            new_files.append(FileItem(
                name=f"{stem}_{sheet_name}.md",
                content=md,
                content_type="text/markdown",
                role=role,
            ))

        wb.close()

        if not new_files:
            return bundle

        bundle.files = new_files
        bundle.context["extension"] = "md"
        return bundle

    @staticmethod
    def _fill_merged_cells(wb: openpyxl.Workbook) -> None:
        for ws in wb.worksheets:
            for merged_range in list(ws.merged_cells.ranges):
                top_left = ws.cell(row=merged_range.min_row, column=merged_range.min_col).value
                ws.unmerge_cells(str(merged_range))
                for row in range(merged_range.min_row, merged_range.max_row + 1):
                    for col in range(merged_range.min_col, merged_range.max_col + 1):
                        ws.cell(row=row, column=col).value = top_left

    def _sheet_to_markdown(self, ws) -> str:
        hidden_rows: set[int] = set()
        hidden_cols: set[int] = set()
        if self._skip_hidden:
            hidden_rows = {r for r, dim in ws.row_dimensions.items() if dim.hidden}
            hidden_cols = {column_index_from_string(c) for c, dim in ws.column_dimensions.items() if dim.hidden}

        rows: list[list[str]] = []
        for row_cells in ws.iter_rows():
            row_num = row_cells[0].row
            if self._skip_hidden and row_num in hidden_rows:
                continue

            values = []
            for cell in row_cells:
                if self._skip_hidden and cell.column in hidden_cols:
                    continue
                values.append(str(cell.value) if cell.value is not None else "")

            if self._skip_empty and all(v == "" for v in values):
                continue

            rows.append(values)

        if not rows:
            return ""

        max_cols = max(len(r) for r in rows)
        for r in rows:
            while len(r) < max_cols:
                r.append("")

        lines = [f"## {ws.title}", ""]
        header = rows[0]
        lines.append("| " + " | ".join(header) + " |")
        lines.append("|" + "|".join("---" for _ in header) + "|")
        for row in rows[1:]:
            lines.append("| " + " | ".join(row) + " |")

        return "\n".join(lines) + "\n"
